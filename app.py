#!/usr/bin/env python3
"""
Rate Deck Automation Tool
Fills competitor rates from Expedia/Booking.com into hotel Rate Deck spreadsheets.
"""
# -*- coding: utf-8 -*-

from flask import Flask, render_template, request, send_file, jsonify
import openpyxl
import csv
from io import BytesIO, StringIO
from datetime import datetime, date
import traceback
import calendar

# Files with many accumulated named styles (Normal, Heading 1, custom, etc.)
# cause openpyxl's apply_stylesheet to spend minutes in NamedStyle._recalculate,
# which indexes every named style's border/font/fill/alignment into the workbook's
# IndexedLists using slow recursive __hash__/__eq__ comparisons.  We never apply
# named styles by name — we only write cell values — so skipping _recalculate is
# safe and eliminates the timeout entirely.  Cell-level xf formatting is loaded
# before this step and is unaffected.
try:
    from openpyxl.styles.named_styles import NamedStyle as _NamedStyle
    _NamedStyle._recalculate = lambda self: None
except Exception:
    pass

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB


# ---------------------------------------------------------------------------
# Competitor mappings
# Each entry: (keyword_in_input_name, keyword_to_find_in_rate_deck_col_A)
# Matching is case-insensitive substring.
# ---------------------------------------------------------------------------

H2O_EXPEDIA_MAP = [
    ('Margaritaville',  'Margaritaville'),
    ('Casa Marina',     'Casa Marina'),
    ('Hyatt Centric',   'Hyatt Centric'),
    ('Ocean Key',       'Ocean Key'),
    ('Pier House',      'Pier House'),
    ('Southernmost',    'Southernmost'),
    ('Reach',           'Reach'),
    ('Courtyard',       'Courtyard'),
]

SMS_EXPEDIA_MAP = [
    ('Margaritaville',  'Margaritaville'),
    ('Casa Marina',     'Casa Marina'),
    ('Hyatt Centric',   'Hyatt Centric'),
    ('Ocean Key',       'Ocean Key'),
    ('Pier House',      'Pier House'),
    ('Southernmost',    'Southernmost'),
    ('Reach',           'Reach'),
    # No Courtyard for SMS
]

# SWM input is always the Lighthouse / Booking.com "Rates" sheet.
SWM_BOOKINGCOM_MAP = [
    ('Southwinds',      'Southwinds'),
    ('Blue Marlin',     'Blue Marlin'),
    ('Best Western',    'Best Western'),
    ('Blue Flamingo',   'Blue Flamingo'),
    ('Courtyard',       'Courtyard'),
    ('Fairfield',       'Fairfield'),
]


# ---------------------------------------------------------------------------
# Value normalisation
# ---------------------------------------------------------------------------

def _to_int_if_whole(v):
    try:
        if v == int(v):
            return int(v)
    except Exception:
        pass
    return v

def normalize_expedia(val):
    """Map Expedia cell to Rate Deck value. Returns None to skip writing."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return _to_int_if_whole(val)
    s = str(val).strip()
    if s in ('S', 'I'):
        return 'SOLD'
    if s == 'M':
        return 'M'
    if s == '-':
        return None   # no data - do not overwrite existing value
    try:
        return _to_int_if_whole(float(s.replace(',', '')))
    except ValueError:
        return s if s else None

def normalize_bookingcom(val):
    """Map Booking.com cell to Rate Deck value. Returns None to skip writing."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return _to_int_if_whole(val)
    s = str(val).strip()
    sl = s.lower()
    if sl == 'sold out':
        return 'SOLD'
    if sl in ('no flex', '--', ''):
        return None   # no data - do not overwrite existing value
    if sl.startswith('los'):
        return s.upper()
    try:
        return _to_int_if_whole(float(s.replace(',', '')))
    except ValueError:
        return s if s else None


# ---------------------------------------------------------------------------
# Sheet / cell lookups
# ---------------------------------------------------------------------------

def _normalise_month_str(s):
    """
    Normalise month strings so strptime can parse them regardless of
    how the spreadsheet spells the month name.
    Handles: full names, 3-letter abbreviations, and 'Sept' (4-letter variant).
    Works for any year.
    """
    s = s.strip()
    # 'Sept YYYY' -> 'Sep YYYY'  (strptime only knows 3-letter 'Sep')
    import re
    s = re.sub(r'\bSept\b', 'Sep', s, flags=re.IGNORECASE)
    return s

def parse_sheet_month_year(sheet_name):
    """
    Return (year, month) from any tab name like:
      'May 2026', 'Jan 2026', 'Sept 2026', 'September 2026', 'Jan 2028', etc.
    Returns None if the name cannot be parsed as a month+year.
    Works for any future year automatically.
    """
    name = _normalise_month_str(sheet_name)
    for fmt in ('%B %Y', '%b %Y'):
        try:
            dt = datetime.strptime(name, fmt)
            return (dt.year, dt.month)
        except ValueError:
            pass
    return None

def find_sheet_for_date(wb, target_date):
    """Return the worksheet whose tab name matches target_date month/year, or None."""
    key = (target_date.year, target_date.month)
    for name in wb.sheetnames:
        if parse_sheet_month_year(name) == key:
            return wb[name]
    return None

def find_col_for_date(ws, target_date, header_row=3, min_col=2):
    """
    Return the column index for target_date in a Rate Deck sheet.
    Strategy 1: direct match on row 3 (works when cells hold actual date values).
    Strategy 2: offset from A4 (works when A4 holds a real datetime anchor).
    Strategy 3: offset from the 1st of the month (fallback when A4 is a formula).
    Pass a data_only-loaded worksheet for best results.
    """
    # Strategy 1: scan row 3 for a real date value
    for col in range(min_col, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val is None or isinstance(val, str):
            continue
        cell_date = val.date() if isinstance(val, datetime) else val
        if isinstance(cell_date, date) and cell_date == target_date:
            return col

    # Strategy 2: A4 contains the 1st of the month as a real datetime
    a4 = ws.cell(4, 1).value
    if isinstance(a4, datetime):
        start = a4.date()
        offset = (target_date - start).days
        if 0 <= offset <= 30:
            col = min_col + offset
            if col <= ws.max_column:
                return col

    # Strategy 3: derive start from target_date's own month (works even when A4 is a formula)
    first_of_month = date(target_date.year, target_date.month, 1)
    offset = (target_date - first_of_month).days
    col = min_col + offset
    if col <= ws.max_column:
        return col

    return None

def find_row_for_label(ws, keyword, search_col=1, min_row=20, max_row=50):
    """Return row where search_col contains keyword (case-insensitive partial match)."""
    kw = keyword.lower()
    for row in range(min_row, max_row + 1):
        val = ws.cell(row, search_col).value
        if val and kw in str(val).lower():
            return row
    return None


# ---------------------------------------------------------------------------
# Expedia date->column map
# ---------------------------------------------------------------------------

def build_expedia_date_col_map(ws):
    """
    Returns {date: col_idx} by reading month headers from row 9
    and day numbers from row 11.
    """
    date_map = {}
    current_month = None
    current_year = None

    for col in range(2, ws.max_column + 1):
        month_cell = ws.cell(9, col).value
        if month_cell and isinstance(month_cell, str) and len(month_cell.strip()) > 4:
            normalised = _normalise_month_str(month_cell).title()  # e.g. "September 2026"
            for fmt in ('%B %Y', '%b %Y'):
                try:
                    dt = datetime.strptime(normalised, fmt)
                    current_month = dt.month
                    current_year = dt.year
                    break
                except ValueError:
                    pass

        if current_month is None:
            continue

        day_val = ws.cell(11, col).value
        if day_val is not None:
            try:
                full_date = date(current_year, current_month, int(day_val))
                date_map[full_date] = col
            except (ValueError, TypeError):
                pass

    return date_map


# ---------------------------------------------------------------------------
# Booking.com date->row map
# ---------------------------------------------------------------------------

def build_bookingcom_date_row_map(ws):
    """Returns {date: row_idx} from the Rates sheet (col C = date, starts row 6)."""
    date_map = {}
    for row in range(6, ws.max_row + 1):
        val = ws.cell(row, 3).value
        if val is None:
            continue
        cell_date = val.date() if isinstance(val, datetime) else val
        if isinstance(cell_date, date):
            date_map[cell_date] = row
    return date_map


# ---------------------------------------------------------------------------
# Core processors
# ---------------------------------------------------------------------------

def process_expedia(master_wb_ro, input_wb, competitor_map, log):
    """
    Compute all cell writes needed from the Expedia input.
    Returns {sheet_name: {(row, col): value}} — does not touch master_wb.
    """
    ws_expedia = input_wb.active  # "Expedia - Revenue management"

    # Scan the whole sheet — hotels may appear in different rows depending on
    # how many summary/header rows Expedia adds.  Skip the user's-own and
    # "Competitive set average rates" rows by matching only on competitor kw.
    expedia_row_for = {}
    last_row = max(ws_expedia.max_row or 0, 100)
    for row in range(5, last_row + 1):
        name = ws_expedia.cell(row, 1).value
        if not name:
            continue
        name_lc = str(name).lower()
        if 'competitive set' in name_lc:
            continue
        for (exp_kw, deck_kw) in competitor_map:
            if exp_kw.lower() in name_lc and deck_kw not in expedia_row_for:
                expedia_row_for[deck_kw] = row
                break

    log.append('Expedia competitors matched: ' + str(list(expedia_row_for.keys())))

    expedia_date_col = build_expedia_date_col_map(ws_expedia)
    if expedia_date_col:
        log.append('Expedia date range: ' + min(expedia_date_col).isoformat() +
                   ' to ' + max(expedia_date_col).isoformat())

    writes = {}
    sheets_missed = 0
    cols_missed = 0
    for target_date, exp_col in expedia_date_col.items():
        deck_ws_ro = find_sheet_for_date(master_wb_ro, target_date)
        if deck_ws_ro is None:
            sheets_missed += 1
            continue
        deck_col = find_col_for_date(deck_ws_ro, target_date)
        if deck_col is None:
            cols_missed += 1
            log.append('No col for ' + target_date.isoformat() + ' in ' + deck_ws_ro.title)
            continue

        sheet_writes = writes.setdefault(deck_ws_ro.title, {})
        for deck_kw, exp_row in expedia_row_for.items():
            deck_row = find_row_for_label(deck_ws_ro, deck_kw)
            if deck_row is None:
                continue
            val = normalize_expedia(ws_expedia.cell(exp_row, exp_col).value)
            if val is not None:
                sheet_writes[(deck_row, deck_col)] = val

    if sheets_missed:
        log.append('Dates skipped (no matching sheet): ' + str(sheets_missed))
    if cols_missed:
        log.append('Dates skipped (no matching column): ' + str(cols_missed))
    return writes


def process_bookingcom(master_wb_ro, input_wb, log):
    """
    Compute all cell writes needed from the Booking.com input.
    Returns {sheet_name: {(row, col): value}} — does not touch master_wb.
    """
    if 'Rates' not in input_wb.sheetnames:
        raise ValueError(
            'Expected a Lighthouse "Rates" sheet but found: ' + str(input_wb.sheetnames) +
            '. Please upload the correct SWM Lighthouse export file.'
        )
    ws_rates = input_wb['Rates']

    bookingcom_col_for = {}
    for col in range(4, ws_rates.max_column + 1):
        header = ws_rates.cell(5, col).value
        if not header:
            continue
        header_lc = str(header).lower()
        for (bc_kw, deck_kw) in SWM_BOOKINGCOM_MAP:
            if bc_kw.lower() in header_lc:
                bookingcom_col_for[deck_kw] = col
                break

    log.append('Booking.com competitors matched: ' + str(list(bookingcom_col_for.keys())))

    bc_date_row = build_bookingcom_date_row_map(ws_rates)
    if bc_date_row:
        log.append('Booking.com date range: ' + min(bc_date_row).isoformat() +
                   ' to ' + max(bc_date_row).isoformat())

    writes = {}
    cols_missed = 0
    for target_date, bc_row in bc_date_row.items():
        deck_ws_ro = find_sheet_for_date(master_wb_ro, target_date)
        if deck_ws_ro is None:
            continue
        deck_col = find_col_for_date(deck_ws_ro, target_date)
        if deck_col is None:
            cols_missed += 1
            log.append('No col for ' + target_date.isoformat() + ' in ' + deck_ws_ro.title)
            continue

        sheet_writes = writes.setdefault(deck_ws_ro.title, {})
        for deck_kw, bc_col in bookingcom_col_for.items():
            deck_row = find_row_for_label(deck_ws_ro, deck_kw)
            if deck_row is None:
                continue
            val = normalize_bookingcom(ws_rates.cell(bc_row, bc_col).value)
            if val is not None:
                sheet_writes[(deck_row, deck_col)] = val

    if cols_missed:
        log.append('Dates skipped (no matching column): ' + str(cols_missed))
    return writes


def apply_writes(master_wb, writes, log):
    """Apply a writes dict to master_wb.  None values clear the cell."""
    cells_written = 0
    for sheet_name, cell_writes in writes.items():
        ws = master_wb[sheet_name]
        for (row, col), val in cell_writes.items():
            ws.cell(row, col).value = val
            cells_written += 1
    log.append('Cells written: ' + str(cells_written))


# ---------------------------------------------------------------------------
# Forecasted Revenue
# ---------------------------------------------------------------------------

def _parse_num(s, force_int=False):
    """Parse a possibly-comma-formatted number string.  Returns None on failure."""
    s = str(s).strip().replace(',', '') if s else ''
    if not s:
        return None
    try:
        v = float(s)
        return int(v) if force_int else _to_int_if_whole(v)
    except ValueError:
        return None


def parse_forecast_csv(csv_bytes):
    """
    Parse a Room Master Report Forecasted Revenue CSV.
    Expected columns (1-indexed):
      A: Revenue Period  e.g. "05/01/2026   Friday"
      B: Daily Filled
      D: Booked Rooms
      F: Cur ADR
    Header is row 1; data starts row 2.
    Returns {date: (daily_filled, booked_rooms, cur_adr)}.
    """
    text = csv_bytes.decode('utf-8-sig', errors='replace')
    reader = csv.reader(StringIO(text))
    data = {}
    first = True
    for row in reader:
        if first:
            first = False
            continue  # skip header row
        if not row or not row[0].strip():
            continue
        date_str = row[0].strip().split()[0]  # "05/01/2026" from "05/01/2026   Friday"
        try:
            d = datetime.strptime(date_str, '%m/%d/%Y').date()
        except ValueError:
            continue
        daily_filled = _parse_num(row[1] if len(row) > 1 else '', force_int=True)
        booked_rooms = _parse_num(row[3] if len(row) > 3 else '', force_int=True)
        cur_adr      = _parse_num(row[5] if len(row) > 5 else '')
        data[d] = (daily_filled, booked_rooms, cur_adr)
    if not data:
        raise ValueError(
            'No valid date rows found in the Forecasted Revenue file. '
            'Expected a CSV with dates in column A (MM/DD/YYYY format), '
            'Daily Filled in column B, Booked Rooms in column D, and ADR in column F.'
        )
    return data


def process_forecast(master_wb_ro, forecast_data, log):
    """
    Compute all cell changes for the weekly forecast update.
    For each affected month sheet:
      Rooms section  (rows 7-11):
        • Row 9 (Rooms Filled)  → value-paste to row 10 (Total Last Week)
        • Rows 8 & 9 cleared
        • Row 9 ← forecast col B (Daily Filled)
        • Row 8 ← forecast col D (Booked Rooms)
      ADR section (rows 14-18):
        • Rows 14-17 → value-paste down one row (14→15, 15→16, 16→17, 17→18)
          overwriting old row 18 (4 Week Prior — deleted)
        • Row 14 cleared then ← forecast col F (Cur ADR)
    Returns {sheet_name: {(row, col): value}} suitable for apply_writes.
    """
    sheet_date_cols = {}  # sheet_name → {date: col}
    skipped = 0
    for d in forecast_data:
        ws_ro = find_sheet_for_date(master_wb_ro, d)
        if ws_ro is None:
            skipped += 1
            continue
        col = find_col_for_date(ws_ro, d)
        if col is None:
            skipped += 1
            continue
        sheet_date_cols.setdefault(ws_ro.title, {})[d] = col

    writes = {}
    for sheet_name, date_col_map in sheet_date_cols.items():
        ws_ro = master_wb_ro[sheet_name]
        # Only operate on the date-column range.  Columns to the right hold
        # formulas (weekly averages, monthly totals) that must be preserved.
        ym = parse_sheet_month_year(sheet_name)
        if ym is None:
            continue
        last_date_col = 1 + calendar.monthrange(ym[0], ym[1])[1]  # col B = day 1
        cw = {}

        # ── Rooms section ──────────────────────────────────────────────────
        # Step 1: copy row 9 → row 10 (date columns only)
        for col in range(2, last_date_col + 1):
            cw[(10, col)] = ws_ro.cell(9, col).value
        # Step 2: clear rows 8 and 9 in date columns (forecast fills them back in)
        for col in range(2, last_date_col + 1):
            cw[(8, col)] = None
            cw[(9, col)] = None
        # Step 3: write new forecast data; these override the clears for matched cols
        for d, col in date_col_map.items():
            daily_filled, booked_rooms, _ = forecast_data[d]
            if daily_filled is not None:
                cw[(9, col)] = daily_filled
            if booked_rooms is not None:
                cw[(8, col)] = booked_rooms

        # ── ADR section ────────────────────────────────────────────────────
        # Step 4: shift rows 14→15, 15→16, 16→17, 17→18 (date columns only)
        for col in range(2, last_date_col + 1):
            cw[(18, col)] = ws_ro.cell(17, col).value
            cw[(17, col)] = ws_ro.cell(16, col).value
            cw[(16, col)] = ws_ro.cell(15, col).value
            cw[(15, col)] = ws_ro.cell(14, col).value
        # Step 5: clear row 14, then write new ADR for matched date columns
        for col in range(2, last_date_col + 1):
            cw[(14, col)] = None
        for d, col in date_col_map.items():
            _, _, cur_adr = forecast_data[d]
            if cur_adr is not None:
                cw[(14, col)] = cur_adr

        writes[sheet_name] = cw

    if skipped:
        log.append('Forecast: ' + str(skipped) + ' dates skipped (no matching sheet/column)')
    total_dates = sum(len(v) for v in sheet_date_cols.values())
    log.append('Forecast: ' + str(len(sheet_date_cols)) + ' sheets updated, ' +
               str(total_dates) + ' dates applied')
    return writes


# ---------------------------------------------------------------------------
# Monthly Occupancy (Rooms Available)
# ---------------------------------------------------------------------------

def parse_occupancy_csv(csv_bytes):
    """
    Parse a Monthly Occupancy (Rooms Available) CSV.
    Expected format:
      Row 1: header — "Month", "", "Day 01", "Day 02", ..., "Day 31", "Available"
      Each month block has 5 rows: GFPL, GFPS, SFB, *TOTAL, * Occ %
      Col A: "Month YYYY" (only on first row of each block; empty for the rest)
      Col B: label
      Cols C-AG (0-indexed 2-32): Day 01 through Day 31
    Rows starting with '*' are skipped (totals / occ %).
    Returns {(year, month): {label: {day: value}}} where day is 1-based.
    """
    text = csv_bytes.decode('utf-8-sig', errors='replace')
    reader = csv.reader(StringIO(text))
    data = {}
    current_key = None
    header_skipped = False
    for row in reader:
        if not header_skipped:
            header_skipped = True
            continue
        if not row:
            continue
        col_a = row[0].strip() if len(row) > 0 else ''
        col_b = row[1].strip() if len(row) > 1 else ''
        if col_a:
            try:
                dt = datetime.strptime(col_a, '%B %Y')
                current_key = (dt.year, dt.month)
            except ValueError:
                current_key = None
        if current_key is None or not col_b or col_b.startswith('*'):
            continue
        day_vals = {}
        for day in range(1, 32):
            csv_idx = day + 1  # col C = day 1 = 0-indexed 2
            raw = row[csv_idx].strip() if len(row) > csv_idx else ''
            val = _parse_num(raw, force_int=True)
            if val is not None:
                day_vals[day] = val
        if day_vals:
            data.setdefault(current_key, {})[col_b] = day_vals
    if not data:
        raise ValueError(
            'No valid monthly occupancy rows found. '
            'Expected a CSV with "Month YYYY" in column A, label in column B, '
            'and day values in columns C onward.'
        )
    return data


def process_occupancy(master_wb_ro, occupancy_data, log):
    """
    Compute cell writes for Monthly Occupancy (Rooms Available).
    For each month/label, finds the destination row using label-based matching
    (rows 40-55, col A) then writes each day value at col = day + 1 (B = day 1).
    Returns {sheet_name: {(row, col): value}}.
    """
    writes = {}
    sheets_matched = 0
    cells_written = 0
    for (year, month), label_days in occupancy_data.items():
        ws_ro = find_sheet_for_date(master_wb_ro, date(year, month, 1))
        if ws_ro is None:
            log.append('Occupancy: no sheet for ' + str(year) + '-' + str(month).zfill(2))
            continue
        sheets_matched += 1
        sheet_writes = writes.setdefault(ws_ro.title, {})
        days_in_month = calendar.monthrange(year, month)[1]
        for label, day_vals in label_days.items():
            dest_row = find_row_for_label(ws_ro, label, search_col=1, min_row=35, max_row=55)
            if dest_row is None:
                log.append('Occupancy: label "' + label + '" not found in ' + ws_ro.title)
                continue
            for day, val in day_vals.items():
                if day > days_in_month:
                    continue  # skip phantom days (e.g. day 29-31 in February)
                sheet_writes[(dest_row, day + 1)] = val
                cells_written += 1
    log.append('Occupancy: ' + str(sheets_matched) + ' sheets updated, ' +
               str(cells_written) + ' cells written')
    return writes


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/generate', methods=['POST'])
def generate():
    prop = request.form.get('property', '').lower()
    initials = request.form.get('initials', '').strip()
    input_file = request.files.get('input_file')
    master_file = request.files.get('master_file')

    if not all([prop, initials, input_file, master_file]):
        return jsonify({'error': 'All fields are required.'}), 400
    if prop not in ('h2o', 'sms', 'swm'):
        return jsonify({'error': 'Unknown property type.'}), 400

    log = []
    try:
        log.append('Loading input: ' + input_file.filename)
        input_wb = openpyxl.load_workbook(BytesIO(input_file.read()), data_only=True)

        # Parse optional forecast CSV (small, negligible memory)
        forecast_file = request.files.get('forecast_file')
        forecast_data = {}
        if forecast_file and forecast_file.filename:
            log.append('Loading forecast: ' + forecast_file.filename)
            forecast_data = parse_forecast_csv(forecast_file.read())
            log.append('Forecast dates loaded: ' + str(len(forecast_data)))

        # Parse optional occupancy CSV
        occupancy_file = request.files.get('occupancy_file')
        occupancy_data = {}
        if occupancy_file and occupancy_file.filename:
            log.append('Loading occupancy: ' + occupancy_file.filename)
            occupancy_data = parse_occupancy_csv(occupancy_file.read())
            log.append('Occupancy months loaded: ' + str(len(occupancy_data)))

        log.append('Loading master: ' + master_file.filename)
        master_bytes = master_file.read()

        # Phase 1: compute all writes using the data_only workbook, then free it
        # before loading the write copy.  Only one full workbook in memory at a time.
        master_wb_ro = openpyxl.load_workbook(BytesIO(master_bytes), data_only=True)
        if prop == 'h2o':
            writes = process_expedia(master_wb_ro, input_wb, H2O_EXPEDIA_MAP, log)
        elif prop == 'sms':
            writes = process_expedia(master_wb_ro, input_wb, SMS_EXPEDIA_MAP, log)
        else:
            writes = process_bookingcom(master_wb_ro, input_wb, log)

        compset_cells = sum(len(v) for v in writes.values())

        forecast_writes = {}
        if forecast_data:
            forecast_writes = process_forecast(master_wb_ro, forecast_data, log)
            for sheet, cw in forecast_writes.items():
                writes.setdefault(sheet, {}).update(cw)

        occupancy_writes = {}
        if occupancy_data:
            occupancy_writes = process_occupancy(master_wb_ro, occupancy_data, log)
            for sheet, cw in occupancy_writes.items():
                writes.setdefault(sheet, {}).update(cw)

        master_wb_ro.close()
        del master_wb_ro

        # Phase 2: load write copy, apply the pre-computed writes, save
        master_wb = openpyxl.load_workbook(BytesIO(master_bytes))
        apply_writes(master_wb, writes, log)

        today = datetime.now().strftime('%y%m%d')
        prop_label = {'h2o': 'H2O', 'sms': 'SMS', 'swm': 'SWM'}[prop]
        filename = today + '-KO-' + prop_label + '-Rate Deck-' + initials + '.xlsx'

        output = BytesIO()
        master_wb.save(output)
        output.seek(0)

        try:
            print('\n'.join(log))
        except Exception:
            pass  # console encoding issues must not abort a good response

        def _cl(provided, cells):
            if not provided:
                return 'skipped'
            return 'done' if cells > 0 else 'failed'

        forecast_cells = sum(len(v) for v in forecast_writes.values())
        occupancy_cells = sum(len(v) for v in occupancy_writes.values())

        response = send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )
        response.headers['X-Checklist-Compset']  = 'done' if compset_cells > 0 else 'failed'
        response.headers['X-Checklist-Rooms']     = _cl(bool(forecast_data), forecast_cells)
        response.headers['X-Checklist-ADR']       = _cl(bool(forecast_data), forecast_cells)
        response.headers['X-Checklist-Occupancy'] = _cl(bool(occupancy_data), occupancy_cells)
        return response

    except Exception as e:
        try:
            traceback.print_exc()
        except Exception:
            pass
        err_msg = str(e).encode('ascii', errors='replace').decode('ascii')
        return jsonify({'error': err_msg, 'log': log}), 500


if __name__ == '__main__':
    print('Rate Deck Automation Tool')
    print('Open http://localhost:5000 in your browser')
    print()
    app.run(debug=False, port=5000)
